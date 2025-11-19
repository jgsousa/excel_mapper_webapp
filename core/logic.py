
import json
import os
from pathlib import Path
from typing import Dict, List, Any
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Environment loader

def load_env():
    try:
        from dotenv import load_dotenv as _load
        _load()
    except Exception:
        pass

# ---------- JSON helpers ----------

def _read_json(path: Path, default):
    if not path.exists():
        return default
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)

def _write_json(path: Path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ---------- Target parsing ----------

def read_target_headers(target_path: Path) -> Dict[str, Dict[str, str]]:
    if not target_path.exists():
        raise FileNotFoundError(f"Target file not found: {target_path}")
    wb = load_workbook(target_path, read_only=True, data_only=True)
    headers_by_sheet: Dict[str, Dict[str, str]] = {}
    for ws in wb.worksheets:
        if ws.title in {"Introduction", "Field List"}:
            continue
        mapping: Dict[str, str] = {}
        row = ws[8]
        for cell in row:
            raw = cell.value
            if raw is None:
                continue
            text = str(raw)
            first_line = text.splitlines()[0].strip()
            if first_line:
                mapping[first_line] = get_column_letter(cell.column)
        headers_by_sheet[ws.title] = mapping
    wb.close()
    return headers_by_sheet


def extract_keys(target_path: Path) -> Dict[str, List[str]]:
    if not target_path.exists():
        raise FileNotFoundError(f"Target file not found: {target_path}")
    wb = load_workbook(target_path, read_only=False, data_only=True)
    keys_by_sheet: Dict[str, List[str]] = {}
    for ws in wb.worksheets:
        if ws.title in {"Introduction", "Field List"}:
            continue
        a7 = ws.cell(row=7, column=1).coordinate
        key_cols: List[str] = []
        found_range = None
        for mr in ws.merged_cells.ranges:
            if a7 in mr:
                found_range = mr
                break
        if found_range:
            min_col = found_range.min_col
            max_col = found_range.max_col
            key_cols = [get_column_letter(c) for c in range(min_col, max_col + 1)]
        else:
            key_cols = ["A"]
        keys_by_sheet[ws.title] = key_cols
    wb.close()
    return keys_by_sheet

# ---------- Source headers ----------

def read_source_headers(xlsx_path: Path) -> List[Dict[str, str]]:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Source file not found: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        val = cell.value
        if val is None:
            continue
        txt = str(val).strip()
        if not txt:
            continue
        headers.append({"header": txt, "column": get_column_letter(cell.column)})
    wb.close()
    return headers

# ---------- Mapping with OpenAI ----------

def _call_openai_for_mapping(source_headers: List[str], target_headers: List[str]) -> List[dict]:
    api_key = os.getenv('OPENAI_API_KEY')
    if not api_key:
        raise RuntimeError('OPENAI_API_KEY missing. Set it in environment or .env file.')
    try:
        from openai import OpenAI
    except Exception as e:
        raise RuntimeError("openai SDK not available. Install 'openai' >= 1.0.0") from e

    client = OpenAI(api_key=api_key)

    sys_prompt = (
        'You are a precise data mapping assistant. You match source column headers to target column headers. '
        'Only return pairs you are highly confident about.'
    )

    lines = [
        'Task: Match source fields to target fields.',
        f'Source headers: {source_headers}',
        f'Target headers: {target_headers}',
        '',
        'Return the result in JSON format [{"source": "...", "target": "..."}]. '
        'No smalltalk no backticks. Use only high confidence matches.'
    ]
    prompt = "".join(lines)

    resp = client.responses.create(
        model='gpt-4.1-mini',
        input=f"{sys_prompt}{prompt}",
        temperature=0.2,
    )

    text = None
    try:
        text = resp.output_text
    except Exception:
        try:
            text = resp.output[0].content[0].text
        except Exception:
            pass
    if not text:
        raise RuntimeError('Empty response from OpenAI')

    text = text.strip()
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        start = text.find('[')
        end = text.rfind(']')
        if start != -1 and end != -1 and end > start:
            data = json.loads(text[start:end+1])
        else:
            raise

    if not isinstance(data, list):
        raise RuntimeError('OpenAI result must be a JSON array of {source, target} objects')

    cleaned = []
    for item in data:
        if not isinstance(item, dict):
            continue
        s = str(item.get('source', '')).strip()
        t = str(item.get('target', '')).strip()
        if s and t:
            cleaned.append({'source': s, 'target': t})
    return cleaned


def run_mapping_for_pair(
    sources_dir: Path,
    filename: str,
    sheet: str,
    target_headers_by_sheet: Dict[str, Dict[str, str]],
    mapping_path: Path,
    headers_cache_path: Path,
) -> List[Dict[str, str]]:
    """
    Returns mapping list for filename::sheet. Uses persisted mapping if present. Otherwise calls OpenAI and persists.
    Output items: {"source": <src_col>, "sourceDescription": <src_name>, "target": <tgt_col>, "targetDescription": <tgt_name>}
    """
    persisted = _read_json(mapping_path, {})
    map_key = f"{filename}::{sheet}"
    if map_key in persisted and isinstance(persisted[map_key], list) and len(persisted[map_key]) > 0:
        return persisted[map_key]

    # Load / compute source headers for this file
    src_path = sources_dir / filename
    if not src_path.exists():
        raise FileNotFoundError(f"Source file not found: {src_path}")

    headers_cache = _read_json(headers_cache_path, {})
    src_hdr_list = headers_cache.get(filename)
    if not src_hdr_list:
        src_hdr_list = read_source_headers(src_path)
        headers_cache[filename] = src_hdr_list
        _write_json(headers_cache_path, headers_cache)

    source_name_to_col = {h['header']: h['column'] for h in src_hdr_list}

    target_map = target_headers_by_sheet.get(sheet, {})
    source_names = list(source_name_to_col.keys())
    target_names = list(target_map.keys())

    if not source_names or not target_names:
        persisted[map_key] = []
        _write_json(mapping_path, persisted)
        return []

    # Call OpenAI to get pairs by names
    name_pairs = _call_openai_for_mapping(source_names, target_names)

    mapped = []
    for pair in name_pairs:
        s_name = pair['source']
        t_name = pair['target']
        s_col = source_name_to_col.get(s_name)
        t_col = target_map.get(t_name)
        if s_col and t_col:
            mapped.append({
                'source': s_col,
                'sourceDescription': s_name,
                'target': t_col,
                'targetDescription': t_name,
            })

    persisted[map_key] = mapped
    _write_json(mapping_path, persisted)
    return mapped


def get_mapping_pair(mapping_path: Path, filename: str, sheet: str) -> List[Dict[str, str]]:
    persisted = _read_json(mapping_path, {})
    return persisted.get(f"{filename}::{sheet}", [])


def set_mapping_pair(mapping_path: Path, filename: str, sheet: str, rows: List[Dict[str, str]]):
    persisted = _read_json(mapping_path, {})
    # simple server-side dedupe by target column to reduce conflicts
    seen_targets = set()
    deduped = []
    for r in rows:
        t = r.get('target')
        if not t or t in seen_targets:
            continue
        seen_targets.add(t)
        deduped.append(r)
    persisted[f"{filename}::{sheet}"] = deduped
    _write_json(mapping_path, persisted)


def delete_mapping_pair(mapping_path: Path, filename: str, sheet: str):
    persisted = _read_json(mapping_path, {})
    key = f"{filename}::{sheet}"
    if key in persisted:
        del persisted[key]
        _write_json(mapping_path, persisted)
