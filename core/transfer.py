
from pathlib import Path
from typing import Dict, List, Any, Tuple
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def _collect_existing_keys(ws, key_cols: List[str], start_row: int) -> Dict[Tuple[Any, ...], int]:
    existing = {}
    if not key_cols:
        return existing
    key_idx = [column_index_from_string(c) - 1 for c in key_cols]  # 0-based for row tuple indexing
    for r, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        tup = tuple(row[i] for i in key_idx)
        existing[tup] = r
    return existing


def _values_for_key_tuple(row_values_by_target_col: Dict[str, Any], key_cols: List[str]) -> Tuple[Any, ...]:
    return tuple(row_values_by_target_col.get(k) for k in key_cols)


def transfer_data(
    config: List[Dict[str, Any]],
    sources_dir: Path,
    target_path: Path,
    mappings: Dict[str, List[Dict[str, str]]],
    keys_by_sheet: Dict[str, List[str]],
    start_row: int = 9,
):
    report = {
        'missing_sources': [],
        'processed': [],
        'updates': 0,
        'inserts': 0,
        'warnings': [],
    }

    if not target_path.exists():
        raise FileNotFoundError(f"Target file not found: {target_path}")

    wb_tgt = load_workbook(target_path)
    ws_map = {ws.title: ws for ws in wb_tgt.worksheets}

    for elem in config:
        filename = elem['filename']
        src_path = sources_dir / filename

        if not src_path.exists():
            report['missing_sources'].append(str(filename))
            continue

        wb_src = load_workbook(src_path, read_only=True, data_only=True)
        ws_src = wb_src.active
        src_max_row = ws_src.max_row or 1

        total_rows = max(0, src_max_row - 1)
        print(f"\nProcessing '{filename}' → Total rows: {total_rows}")

        # Pre-read all source rows into memory (row 1 = header, row 2+ = data)
        all_src_rows = list(ws_src.iter_rows(min_row=1, values_only=True))
        wb_src.close()

        if len(all_src_rows) < 2:
            continue

        # Build column-name → 0-based index map from header row
        header = all_src_rows[0]
        src_col_idx: Dict[str, int] = {}
        for col_letter in set(
            m['source']
            for tgt in elem['target']
            for map_key in [f"{filename}::{tgt['sheet']}"]
            for m in mappings.get(map_key, [])
        ):
            ci = column_index_from_string(col_letter) - 1
            src_col_idx[col_letter] = ci

        processed_count = 0

        for tgt in elem['target']:
            sheet = tgt['sheet']
            ws_tgt = ws_map.get(sheet)

            if ws_tgt is None:
                report['warnings'].append(f"Target sheet '{sheet}' not found in Target.xlsx. Skipped.")
                continue

            map_key = f"{filename}::{sheet}"
            map_rows = mappings.get(map_key, [])

            if not map_rows:
                report['warnings'].append(f"No mapping for {map_key}. Skipped.")
                continue

            t_to_s = {m['target']: m['source'] for m in map_rows}
            # Pre-convert target column letters to integer indices once
            t_col_idx = {t: column_index_from_string(t) for t in t_to_s}

            key_cols = keys_by_sheet.get(sheet, ['A']) or ['A']
            existing_index = _collect_existing_keys(ws_tgt, key_cols, start_row)

            # Track max_row manually to avoid repeated ws_tgt.max_row calls
            current_max_row = ws_tgt.max_row or (start_row - 1)

            for data_row in all_src_rows[1:]:
                processed_count += 1

                row_values_by_target_col: Dict[str, Any] = {}
                for t_col, s_col in t_to_s.items():
                    ci = src_col_idx.get(s_col)
                    val = data_row[ci] if ci is not None and ci < len(data_row) else None
                    row_values_by_target_col[t_col] = val

                key_tuple = _values_for_key_tuple(row_values_by_target_col, key_cols)
                missing_key = any(
                    v is None or (isinstance(v, str) and v.strip() == '')
                    for v in key_tuple
                )

                if (not missing_key) and (key_tuple in existing_index):
                    dest_row = existing_index[key_tuple]
                    is_update = True
                else:
                    current_max_row = max(current_max_row + 1, start_row)
                    dest_row = current_max_row
                    existing_index[key_tuple] = dest_row
                    is_update = False

                for t_col, val in row_values_by_target_col.items():
                    ws_tgt.cell(row=dest_row, column=t_col_idx[t_col]).value = val

                if is_update:
                    report['updates'] += 1
                else:
                    report['inserts'] += 1

                if processed_count % 100 == 0 or processed_count == total_rows:
                    print(
                        f"  Progress: {processed_count}/{total_rows} | "
                        f"Updates: {report['updates']} | Inserts: {report['inserts']}"
                    )

            report['processed'].append({
                'source': str(filename),
                'sheet': sheet
            })

    wb_tgt.save(target_path)
    wb_tgt.close()

    print("\n✅ Final Report:")
    print(report)

    return report
